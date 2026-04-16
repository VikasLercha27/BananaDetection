# ============================================================
# hyperparameter_experiments_v2.py
#
# Random Search over hyperparameter combinations.
# Tracks: Train / Val / Test accuracy, Precision, Recall,
#         F1 Score, Parameter Count, Gap Analysis.
# Highlights: best efficiency (low params + high accuracy),
#             gap analysis (train vs val overfit detection).
#
# pip install openpyxl tensorflow scikit-learn numpy joblib
# ============================================================

import os
import random
import numpy as np
import joblib
import datetime
import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout, BatchNormalization, Activation, Input
from tensorflow.keras.regularizers import l1_l2
from tensorflow.keras.callbacks import EarlyStopping, ReduceLROnPlateau
from tensorflow.keras.optimizers import Adam
from sklearn.metrics import precision_score, recall_score, f1_score, classification_report
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────
BACKBONE      = "efficientnet"
FEATURE_DIR   = f"features/{BACKBONE}"
MODEL_DIR     = f"models/{BACKBONE}"
EXCEL_PATH    = "hyperparameter_experiments_v2.xlsx"
INPUT_DIM     = 1287       # your actual EfficientNet feature dimension
NUM_RUNS      = 1000       # total random combinations to try
RANDOM_SEED   = 42
MAX_EPOCHS    = 200
ES_PATIENCE   = 10
LR_PATIENCE   = 5

random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

# ── Hyperparameter search space ────────────────────────────────────────────
SEARCH_SPACE = {
    "neurons_layer1" : [64, 128, 256, 512],
    "neurons_layer2" : [64, 128, 256, 512],
    "activation"     : ["relu", "tanh", "sigmoid", "softmax"],
    "l1_reg"         : [1.0, 0.1, 0.2, 0.3, 0.5, 0.01, 0.03, 0.001, 0.005],
    "l2_reg"         : [1.0, 0.1, 0.2, 0.3, 0.5, 0.01, 0.03, 0.001, 0.005],
    "learning_rate"  : [1.0, 0.1, 0.2, 0.3, 0.5, 0.01, 0.03, 0.001, 0.005],
    "dropout_1"      : [0.2, 0.3, 0.4, 0.5],
    "dropout_2"      : [0.2, 0.3, 0.4, 0.5],
    "batch_size"     : [32, 40, 50, 60, 90],
}

# ── Column definitions ─────────────────────────────────────────────────────
COLUMNS = [
    "Run #",             # A
    "Timestamp",         # B
    "Neurons L1",        # C
    "Neurons L2",        # D
    "Activation",        # E
    "L1 Reg",            # F
    "L2 Reg",            # G
    "Learning Rate",     # H
    "Dropout L1",        # I
    "Dropout L2",        # J
    "Batch Size",        # K
    "Parameters",        # L  ← param count in short form
    "Epochs Run",        # M
    "Train Accuracy",    # N
    "Val Accuracy",      # O
    "Test Accuracy",     # P
    "Train Loss",        # Q
    "Val Loss",          # R
    "Test Loss",         # S
    "Precision",         # T
    "Recall",            # U
    "F1 Score",          # V
    "Val vs Train Gap",  # W  ← gap analysis
    "Efficiency Score",  # X  ← accuracy / log10(params)
    "Notes",             # Y
]

# Column index map (1-based)
COL = {name: idx for idx, name in enumerate(COLUMNS, start=1)}

# ── Colour fills ───────────────────────────────────────────────────────────
FILL = {
    "title"     : PatternFill("solid", start_color="0D2D4E"),
    "header"    : PatternFill("solid", start_color="1F4E79"),
    "subheader" : PatternFill("solid", start_color="2E75B6"),
    "alt_row"   : PatternFill("solid", start_color="EBF3FB"),
    "good"      : PatternFill("solid", start_color="C6EFCE"),   # green
    "warn"      : PatternFill("solid", start_color="FFEB9C"),   # yellow
    "bad"       : PatternFill("solid", start_color="FFC7CE"),   # red
    "gold"      : PatternFill("solid", start_color="FFD700"),   # gold — best efficiency
    "purple"    : PatternFill("solid", start_color="E8D5FF"),   # purple — low param high acc
}

FONT = {
    "title"   : Font(name="Arial", bold=True, size=14, color="FFFFFF"),
    "header"  : Font(name="Arial", bold=True, size=10, color="FFFFFF"),
    "body"    : Font(name="Arial", size=10),
    "bold"    : Font(name="Arial", bold=True, size=10),
    "white"   : Font(name="Arial", bold=True, size=10, color="FFFFFF"),
}

thin = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)

center = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Helpers ────────────────────────────────────────────────────────────────
def compute_params(params):
    n1 = params["neurons_layer1"]
    n2 = params["neurons_layer2"]
    dense1 = (INPUT_DIM * n1) + n1
    bn1    = 4 * n1
    dense2 = (n1 * n2) + n2
    bn2    = 4 * n2
    out    = n2 + 1
    return dense1 + bn1 + dense2 + bn2 + out

def short_param(n):
    if n >= 1_000_000:
        return f"{n/1_000_000:.2f}M"
    return f"{n/1_000:.1f}K"

def colour_accuracy(cell, acc):
    if acc >= 0.97:
        cell.fill = FILL["good"]
    elif acc >= 0.94:
        cell.fill = FILL["warn"]
    else:
        cell.fill = FILL["bad"]

def colour_gap(cell, gap):
    if abs(gap) <= 0.01:
        cell.fill = FILL["good"]
    elif abs(gap) <= 0.03:
        cell.fill = FILL["warn"]
    else:
        cell.fill = FILL["bad"]

def colour_f1(cell, f1):
    if f1 >= 0.97:
        cell.fill = FILL["good"]
    elif f1 >= 0.93:
        cell.fill = FILL["warn"]
    else:
        cell.fill = FILL["bad"]

# ── Random combinations generator ─────────────────────────────────────────
def generate_random_combinations(n):
    combos = []
    seen   = set()
    attempts = 0
    while len(combos) < n and attempts < n * 20:
        attempts += 1
        combo = {k: random.choice(v) for k, v in SEARCH_SPACE.items()}
        key   = tuple(combo[k] for k in sorted(combo))
        if key not in seen:
            seen.add(key)
            combos.append(combo)
    return combos

# ── Model builder ──────────────────────────────────────────────────────────
def build_model(params):
    act = params["activation"]
    reg = l1_l2(l1=params["l1_reg"], l2=params["l2_reg"])

    model = Sequential([
        Input(shape=(INPUT_DIM,)),

        Dense(params["neurons_layer1"], kernel_regularizer=reg),
        BatchNormalization(),
        Activation(act),
        Dropout(params["dropout_1"]),

        Dense(params["neurons_layer2"], kernel_regularizer=reg),
        BatchNormalization(),
        Activation(act),
        Dropout(params["dropout_2"]),

        Dense(1, activation="sigmoid"),
    ])

    model.compile(
        optimizer=Adam(learning_rate=params["learning_rate"]),
        loss="binary_crossentropy",
        metrics=["accuracy"],
    )
    return model

# ── Single experiment ──────────────────────────────────────────────────────
def run_experiment(X_train, y_train, X_val, y_val, X_test, y_test, params):
    tf.keras.backend.clear_session()

    model = build_model(params)

    callbacks = [
        EarlyStopping(monitor="val_loss", patience=ES_PATIENCE,
                      restore_best_weights=True, verbose=0),
        ReduceLROnPlateau(monitor="val_loss", factor=0.5,
                          patience=LR_PATIENCE, min_lr=1e-7, verbose=0),
    ]

    history = model.fit(
        X_train, y_train,
        validation_data=(X_val, y_val),
        epochs=MAX_EPOCHS,
        batch_size=params["batch_size"],
        callbacks=callbacks,
        verbose=0,
    )

    epochs_run = len(history.history["loss"])

    tr  = model.evaluate(X_train, y_train, verbose=0)
    vl  = model.evaluate(X_val,   y_val,   verbose=0)
    ts  = model.evaluate(X_test,  y_test,  verbose=0)

    # Precision / Recall / F1 from sklearn (more reliable than Keras metrics)
    y_pred = (model.predict(X_test, verbose=0) >= 0.5).astype(int).flatten()
    prec   = precision_score(y_test, y_pred, zero_division=0)
    rec    = recall_score(y_test,    y_pred, zero_division=0)
    f1     = f1_score(y_test,        y_pred, zero_division=0)

    return {
        "epochs_run"  : epochs_run,
        "train_loss"  : tr[0],  "train_acc"  : tr[1],
        "val_loss"    : vl[0],  "val_acc"    : vl[1],
        "test_loss"   : ts[0],  "test_acc"   : ts[1],
        "precision"   : prec,
        "recall"      : rec,
        "f1"          : f1,
    }

# ── Excel setup ────────────────────────────────────────────────────────────
def setup_excel(path):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        next_row   = ws.max_row + 1
        run_number = ws.max_row
        print(f"  Appending — next row: {next_row}, run#: {run_number}")
        return wb, ws, next_row, run_number

    wb = Workbook()
    ws = wb.active
    ws.title = "Experiments"

    # Title row
    last_col = get_column_letter(len(COLUMNS))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "Mango Detection — Extended Hyperparameter Search (Random Search v2)"
    c.font      = FONT["title"]
    c.fill      = FILL["title"]
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, name in enumerate(COLUMNS, start=1):
        c            = ws.cell(row=2, column=col_idx, value=name)
        c.font       = FONT["header"]
        c.fill       = FILL["header"]
        c.alignment  = center_wrap
        c.border     = thin
    ws.row_dimensions[2].height = 36

    # Column widths
    widths = [7,18,11,11,11,9,9,14,11,11,11,
              12,11,15,15,15,12,10,10,12,10,10,16,15,20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    print("  Created new Excel file.")
    return wb, ws, 3, 1

# ── Write one row ──────────────────────────────────────────────────────────
def write_row(ws, row, run_num, params, results):
    p_count   = compute_params(params)
    p_short   = short_param(p_count)
    train_acc = results["train_acc"]
    val_acc   = results["val_acc"]
    test_acc  = results["test_acc"]
    gap       = val_acc - train_acc
    # Efficiency: how much accuracy per unit of model complexity
    import math
    efficiency = round(val_acc / math.log10(p_count), 6) if p_count > 0 else 0

    row_data = [
        run_num,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        params["neurons_layer1"],
        params["neurons_layer2"],
        params["activation"],
        params["l1_reg"],
        params["l2_reg"],
        params["learning_rate"],
        params["dropout_1"],
        params["dropout_2"],
        params["batch_size"],
        p_short,
        results["epochs_run"],
        round(train_acc,           4),
        round(val_acc,             4),
        round(test_acc,            4),
        round(results["train_loss"],4),
        round(results["val_loss"],  4),
        round(results["test_loss"], 4),
        round(results["precision"], 4),
        round(results["recall"],    4),
        round(results["f1"],        4),
        round(gap,                  4),
        efficiency,
        "",   # Notes
    ]

    for col_idx, value in enumerate(row_data, start=1):
        c            = ws.cell(row=row, column=col_idx, value=value)
        c.font       = FONT["body"]
        c.alignment  = center
        c.border     = thin
        if row % 2 == 0:
            c.fill = FILL["alt_row"]

    # Colour-code accuracy columns
    colour_accuracy(ws.cell(row=row, column=COL["Train Accuracy"]), train_acc)
    colour_accuracy(ws.cell(row=row, column=COL["Val Accuracy"]),   val_acc)
    colour_accuracy(ws.cell(row=row, column=COL["Test Accuracy"]),  test_acc)

    # Colour-code F1
    colour_f1(ws.cell(row=row, column=COL["F1 Score"]), results["f1"])

    # Colour-code gap
    colour_gap(ws.cell(row=row, column=COL["Val vs Train Gap"]), gap)

    return p_count, val_acc, efficiency

# ── Post-run analysis sheet ────────────────────────────────────────────────
def add_analysis_sheet(wb, data_rows, first_data_row, last_data_row):
    """
    Add a second sheet with:
      1. Best runs table (top 10 by Val Accuracy)
      2. Efficiency highlights (high acc + low params)
      3. Gap analysis summary
    """
    if "Analysis" in wb.sheetnames:
        del wb["Analysis"]

    ws2 = wb.create_sheet("Analysis")

    def hdr(cell_ref, text, fill_key="header"):
        c = ws2[cell_ref]
        c.value     = text
        c.font      = FONT["white"] if fill_key in ("header","title","subheader") else FONT["bold"]
        c.fill      = FILL[fill_key]
        c.alignment = center
        c.border    = thin

    def val(cell_ref, v, fill_key=None):
        c = ws2[cell_ref]
        c.value     = v
        c.font      = FONT["body"]
        c.alignment = center
        c.border    = thin
        if fill_key:
            c.fill = FILL[fill_key]

    # ── Section 1: Top 10 by Val Accuracy ─────────────────────────────────
    ws2.merge_cells("A1:J1")
    hdr("A1", "TOP 10 RUNS — Highest Validation Accuracy", "title")

    headers_top = ["Run #","Neurons L1","Neurons L2","Activation",
                   "Learning Rate","Params","Val Acc","Test Acc","F1 Score","Gap"]
    for ci, h in enumerate(headers_top, start=1):
        hdr(ws2.cell(row=2, column=ci).coordinate, h)

    top10_formula_note = (
        f"=LARGE(Experiments!O{first_data_row}:O{last_data_row},ROW()-2)"
    )
    # We write a note that user can sort — actual top10 filled via INDEX/MATCH
    for rank in range(1, 11):
        r = rank + 2
        # Val accuracy rank formula
        ws2.cell(row=r, column=7).value  = f"=LARGE(Experiments!O$3:O${last_data_row},{rank})"
        ws2.cell(row=r, column=7).font   = FONT["body"]
        ws2.cell(row=r, column=7).alignment = center
        ws2.cell(row=r, column=7).border = thin
        ws2.cell(row=r, column=7).fill   = FILL["good"]

        # Matching columns via INDEX/MATCH on Val Accuracy rank
        match_cols = {
            1: f"=INDEX(Experiments!A$3:A${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            2: f"=INDEX(Experiments!C$3:C${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            3: f"=INDEX(Experiments!D$3:D${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            4: f"=INDEX(Experiments!E$3:E${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            5: f"=INDEX(Experiments!H$3:H${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            6: f"=INDEX(Experiments!L$3:L${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            8: f"=INDEX(Experiments!P$3:P${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            9: f"=INDEX(Experiments!V$3:V${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
            10:f"=INDEX(Experiments!W$3:W${last_data_row},MATCH(G{r},Experiments!O$3:O${last_data_row},0))",
        }
        for col_i, formula in match_cols.items():
            c = ws2.cell(row=r, column=col_i)
            c.value     = formula
            c.font      = FONT["body"]
            c.alignment = center
            c.border    = thin

    # ── Section 2: Best Efficiency (high accuracy, low params) ────────────
    ws2.merge_cells("A14:J14")
    hdr("A14", "EFFICIENCY HIGHLIGHT — Highest Efficiency Score (Val Acc / log10(Params))", "subheader")

    for ci, h in enumerate(headers_top, start=1):
        hdr(ws2.cell(row=15, column=ci).coordinate, h)

    for rank in range(1, 11):
        r = rank + 15
        ws2.cell(row=r, column=8).value  = f"=LARGE(Experiments!X$3:X${last_data_row},{rank})"
        ws2.cell(row=r, column=8).font   = FONT["body"]
        ws2.cell(row=r, column=8).alignment = center
        ws2.cell(row=r, column=8).border = thin
        ws2.cell(row=r, column=8).fill   = FILL["purple"]

        eff_match_cols = {
            1: f"=INDEX(Experiments!A$3:A${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            2: f"=INDEX(Experiments!C$3:C${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            3: f"=INDEX(Experiments!D$3:D${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            4: f"=INDEX(Experiments!E$3:E${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            5: f"=INDEX(Experiments!H$3:H${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            6: f"=INDEX(Experiments!L$3:L${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            7: f"=INDEX(Experiments!O$3:O${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            9: f"=INDEX(Experiments!V$3:V${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
            10:f"=INDEX(Experiments!W$3:W${last_data_row},MATCH(H{r},Experiments!X$3:X${last_data_row},0))",
        }
        for col_i, formula in eff_match_cols.items():
            c = ws2.cell(row=r, column=col_i)
            c.value     = formula
            c.font      = FONT["body"]
            c.alignment = center
            c.border    = thin

    # ── Section 3: Gap Analysis Summary ───────────────────────────────────
    ws2.merge_cells("A27:J27")
    hdr("A27", "GAP ANALYSIS — Val vs Train Accuracy (Overfitting Detection)", "title")

    gap_labels = [
        ("A28", "Metric", "header"),
        ("B28", "Value",  "header"),
        ("A29", "Max Train Accuracy",              None),
        ("B29", f"=MAX(Experiments!N$3:N${last_data_row})", "good"),
        ("A30", "Max Val Accuracy",                None),
        ("B30", f"=MAX(Experiments!O$3:O${last_data_row})", "good"),
        ("A31", "Max Test Accuracy",               None),
        ("B31", f"=MAX(Experiments!P$3:P${last_data_row})", "good"),
        ("A32", "Avg Val vs Train Gap",            None),
        ("B32", f"=AVERAGE(Experiments!W$3:W${last_data_row})", "warn"),
        ("A33", "Min Gap (least overfit run)",     None),
        ("B33", f"=MIN(ABS(Experiments!W$3:W${last_data_row}))", "good"),
        ("A34", "Max Gap (most overfit run)",      None),
        ("B34", f"=MAX(ABS(Experiments!W$3:W${last_data_row}))", "bad"),
        ("A35", "Runs with gap <= 0.01 (tight fit)",None),
        ("B35", f'=COUNTIF(Experiments!W$3:W${last_data_row},">=- 0.01")-COUNTIF(Experiments!W$3:W${last_data_row},">0.01")', "good"),
        ("A36", "Runs with gap > 0.05 (overfit)",  None),
        ("B36", f'=COUNTIF(Experiments!W$3:W${last_data_row},"<-0.05")', "bad"),
        ("A37", "Best F1 Score",                   None),
        ("B37", f"=MAX(Experiments!V$3:V${last_data_row})", "good"),
        ("A38", "Avg F1 Score",                    None),
        ("B38", f"=AVERAGE(Experiments!V$3:V${last_data_row})", "warn"),
    ]

    for ref, text, fill_key in gap_labels:
        c = ws2[ref]
        c.value     = text
        c.font      = FONT["bold"] if fill_key == "header" else FONT["body"]
        c.fill      = FILL[fill_key] if fill_key else PatternFill()
        c.alignment = center
        c.border    = thin
        if fill_key == "header":
            c.font = FONT["white"]
            c.fill = FILL["header"]

    # Column widths for analysis sheet
    for col, w in zip("ABCDEFGHIJ", [10,18,14,14,16,10,12,12,12,14]):
        ws2.column_dimensions[col].width = w

# ── Main ───────────────────────────────────────────────────────────────────
def main():
    print("\n[1/4] Loading features …")
    scaler  = joblib.load(f"{MODEL_DIR}/scaler.pkl")
    X_train = scaler.transform(np.load(f"{FEATURE_DIR}/X_train.npy"))
    y_train = np.load(f"{FEATURE_DIR}/y_train.npy")
    X_val   = scaler.transform(np.load(f"{FEATURE_DIR}/X_val.npy"))
    y_val   = np.load(f"{FEATURE_DIR}/y_val.npy")
    X_test  = scaler.transform(np.load(f"{FEATURE_DIR}/X_test.npy"))
    y_test  = np.load(f"{FEATURE_DIR}/y_test.npy")
    print(f"  Train:{X_train.shape}  Val:{X_val.shape}  Test:{X_test.shape}")
    print(f"  Feature dimension confirmed: {X_train.shape[1]}")

    print(f"\n[2/4] Generating {NUM_RUNS} random combinations …")
    combos = generate_random_combinations(NUM_RUNS)
    print(f"  {len(combos)} unique combinations ready.")

    print(f"\n[3/4] Setting up Excel → {EXCEL_PATH}")
    wb, ws, next_row, run_number = setup_excel(EXCEL_PATH)
    first_data_row = 3

    print(f"\n[4/4] Running experiments …")
    print(f"{'='*70}")

    all_results = []

    for i, params in enumerate(combos, start=1):
        p_short = short_param(compute_params(params))
        print(f"  Run {run_number:>4}/{run_number + len(combos) - i:>4} "
              f"| combo {i:>4}/{len(combos)} "
              f"| params={p_short} "
              f"| act={params['activation']:<8} "
              f"| lr={params['learning_rate']}")

        try:
            results = run_experiment(
                X_train, y_train, X_val, y_val, X_test, y_test, params
            )
            print(f"         → train={results['train_acc']:.4f} "
                  f"val={results['val_acc']:.4f} "
                  f"test={results['test_acc']:.4f} "
                  f"f1={results['f1']:.4f} "
                  f"epochs={results['epochs_run']}")

            p_count, val_acc, efficiency = write_row(
                ws, next_row, run_number, params, results
            )
            all_results.append({
                "row": next_row, "p_count": p_count,
                "val_acc": val_acc, "efficiency": efficiency
            })

            # Save every run — no data lost on crash
            wb.save(EXCEL_PATH)
            next_row   += 1
            run_number += 1

        except Exception as e:
            print(f"         [ERROR] {e} — skipping.")
            continue

    last_data_row = next_row - 1

    # ── Highlight best efficiency rows in Experiments sheet ───────────────
    if all_results:
        # Top 10% by efficiency score
        sorted_eff = sorted(all_results, key=lambda x: x["efficiency"], reverse=True)
        top_eff_rows = {r["row"] for r in sorted_eff[:max(1, len(sorted_eff)//10)]}

        # Highlight Parameters column gold for top efficiency rows
        for r in all_results:
            if r["row"] in top_eff_rows:
                cell = ws.cell(row=r["row"], column=COL["Parameters"])
                cell.fill = FILL["gold"]
                cell.font = Font(name="Arial", bold=True, size=10, color="7B3F00")

        # Summary row at bottom
        summary_row = last_data_row + 2
        last_col    = get_column_letter(len(COLUMNS))
        ws.merge_cells(f"A{summary_row}:{last_col}{summary_row}")
        c = ws[f"A{summary_row}"]
        c.value = (f"Total runs: {run_number-1}  |  "
                   f"Best Val Acc: =TEXT(MAX(O3:O{last_data_row}),\"0.0000\")  |  "
                   f"Best F1: =TEXT(MAX(V3:V{last_data_row}),\"0.0000\")  |  "
                   f"Avg Val Acc: =TEXT(AVERAGE(O3:O{last_data_row}),\"0.0000\")")
        c.font      = FONT["white"]
        c.fill      = FILL["subheader"]
        c.alignment = center

    # ── Add Analysis sheet ─────────────────────────────────────────────────
    print("\n  Building Analysis sheet …")
    add_analysis_sheet(wb, all_results, first_data_row, last_data_row)
    wb.save(EXCEL_PATH)

    print(f"\n{'='*70}")
    print(f"  All {run_number-1} experiments complete.")
    print(f"  Results saved → {EXCEL_PATH}")
    print(f"  Gold highlighted rows = top 10% efficiency (high acc, low params)")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()